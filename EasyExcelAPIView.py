import os, datetime
from django.db import models
from openpyxl import Workbook
from django.http import HttpResponse
from rest_framework.views import APIView
from rest_framework import status
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.drawing.image import Image
# from core.settings import BASE_DIR # Sirve para agregar una imagen al excel


class ExcelResponseMixin(object):
    """
    Mixin para exportar datos a un archivo de Excel 2007 o superior.
    """
    def get_excel_response(self, wb, filename):
        """
        :param wb: Objeto de tipo Workbook
        :param filename: Nombre del archivo a exportar
        """
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',)
        response['Content-Disposition'] = 'attachment; filename=%s' % filename
        response['charset'] = 'utf-8'
        response['status'] = status.HTTP_200_OK
        wb.save(response)
        return response


class CustomWorkbook(ExcelResponseMixin):
    """
    Es una clase que va permitir estilizar, agregar datos, filtrar datos, etc.
    """

    def __init__(self, model, fields, filename=None, sheet_name=None, header_fields=None, custom_func_row=None, *args, **kwargs):
        self.model = model
        self.fields = fields
        self.wb = Workbook()
        self.ws = self.wb.active
        self.title = sheet_name 
        self.filename = filename
        self.ws.title = self.title
        self.header = header_fields
        self.queryset = None
        self.custom_func_row = custom_func_row

    def add_custom_data(self, queryset):
        """
        Filtra los datos a exportar.
        """
        self.queryset = queryset

    def get_body_data(self):
            """
            Obtiene los datos a exportar.
            """
            if self.queryset:
                return self.queryset.values_list(*self.fields)
            return self.model.objects.values_list(*self.fields)

    def set_body_data(self):
        """
        Agrega los datos al worksheet.
        """
        [self.ws.append(self.custom_func_row(row)) for row in self.get_body_data()]

    def set_header(self):
        """
        Agrega los encabezados al worksheet.
        """
        self.ws.append(self.header) 

    """def draw_image(self, row, column):
        #Agrega una imagen al worksheet.
        
        # path static files
        image_path = os.path.join(BASE_DIR, 'static/img/ft_logo_black.jpg')
        img = Image(image_path)
        img.width = 500
        img.height = 200
        self.ws.add_image(img, '%s%s' % (self.get_column_letter(column), row))"""

    def set_adjust_column_width(self):
        """
        Ajusta el ancho de las columnas.
        """
        for column_cells in self.ws.columns:
            length = max(len(str(cell.value)) for cell in column_cells)
            self.ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    def set_style_header(self):
        """
        Estiliza los encabezados de las columnas.
        """
        range_header = 'A1:%s1' % self.get_column_letter(len(self.header))
        for row in self.ws[range_header]:
            for cell in row:
                cell.font = Font(bold=True, color='0a0a1e')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.border = Border(left=Side(border_style='thin', color='0a0a1e'),
                                     right=Side(border_style='thin', color='0a0a1e'),
                                     top=Side(border_style='thin', color='0a0a1e'),
                                     bottom=Side(border_style='thin', color='0a0a1e'))
                cell.fill = PatternFill(start_color='eef7ff', end_color='eef7ff', fill_type='solid')
                
    def get_column_letter(self, index):
        """
        Obtiene la letra de la columna.
        """
        return self.ws.cell(row=1, column=index).column_letter if index > 0 else None 

    def get_worksheet(self):
        """
        Obtiene el worksheet.
        """
        return self.ws

    def get_workbook(self):
        """
        Obtiene el workbook.
        """
        return self.wb

    def get_filename(self):
        """
        Obtiene el nombre del archivo a exportar.
        """
        if self.filename:
            return self.filename
        elif self.title:
            return self.title + '.xlsx'
        return 'Reporte.xlsx'

    def compile(self):
        """
        Compila los datos.
        """
        self.set_header()
        self.set_style_header()
        self.set_body_data()
        self.set_adjust_column_width()

    def get_response(self):
        """
        Obtiene el objeto response.
        """
        self.compile()
        return self.get_excel_response(self.get_workbook(), self.filename)


class EasyExcelAPIView(APIView, ExcelResponseMixin):
    """
    Clase para exportar datos a un archivo de Excel 2007 o superior.
    :param permission_classes: Permisos de acceso a la vista (opcional)
    :param model: Modelo de datos (obligatorio) 
    :param fields: Campos a exportar (opcional), por defecto son todos los campos 
                   del modelo, excluyendo los campos con relaciones. 
    :param filename: Nombre del archivo (opcional) por defecto es el nombre del modelo
    :param sheet_name: Nombre de la hoja (opcional) por defecto es el nombre del modelo
    :param header: Encabezados de las columnas (opcional) por defecto son los nombres de los campos
    :param with_relacion: Exportar campos con relaciones (opcional) es False por defecto.
    :param kwargs_custom_func_row: Diccionario con funciones personalizadas para obtener los datos de la fila (opcional)
    :param header_extra: Encabezados extra de las columnas (opcional), fuera de los campos del modelo

    Ejemplo de uso:
    class ExportarExcelView(EasyExcelAPIView):
        model = Cliente
        fields = ['id', 'nombre', 'apellido', 'email', 'telefono', 'direccion', 'ciudad', 'pais']
        filename = 'clientes.xlsx'
        sheet_name = 'Clientes'
        header = ['ID', 'Nombre', 'Apellido', 'Email', 'Teléfono', 'Dirección', 'Ciudad', 'País']
        header_extra = ['Fecha de creación', 'Fecha de actualización']
        with_relacion = True # Exportar campos con relaciones
        kwargs_custom_func_row = {'extra1_id': get_fecha_creacion, 'extra2_id': get_fecha_actualizacion}
    """

    permission_classes = [] # Permisos de acceso a la vista (optional)
    model = None # Model de Django to export data (required)
    fields = None # Fields to export (optional)
    filename = None  # Name of the file (optional)
    sheet_name = None # Name of the sheet (optional)
    header = None # Header of the columns (optional) parte de los campos del modelo
    header_extra = None # Header extra of the columns (optional), fuera de los campos del modelo
    with_relacion = False # Export fields with relation (optional)
    kwargs_custom_func_row = None # Dictionary with custom functions to get data of the row (optional)

    def __init__(self, *args, **kwargs):
        super(EasyExcelAPIView, self).__init__(*args, **kwargs)
        self.fields = self.get_fields()
        self.filename = self.get_filename()
        self.sheet_name = self.get_sheet_name()
        self.header = self.get_header()
        self.kwargs_custom_func_row = self.get_kwargs_custom_func_row()
        self.__customs_func_row = self.get_custom_func_row(**self.kwargs_custom_func_row) # Custom functions to get data of the row
  
    def get(self, request, format=None, *args, **kwargs):
        """
        Obtiene los datos a exportar.
        """
        custom_wb = CustomWorkbook(self.model, self.fields, self.filename, self.sheet_name,\
                                   self.header, self.__customs_func_row)
        self.filter_data_wb(custom_wb, *args, **self.get_kwargs_filter_request(request))
        return custom_wb.get_response()

    def filter_data_wb(self,  custom_wb, *args, **kwargs):
        """
        Filtra los datos a exportar.
        """
        if kwargs:
            custom_wb.add_custom_data(self.model.objects.filter(**kwargs))
        
    def get_kwargs_filter_request(self, request):
        """
        Obtiene los parámetros de la petición.
        """
        kwargs = {}
        for key, value in request.GET.items():
            #if key in self.get_fields():
            kwargs[key] = value
        return kwargs

    def get_fields(self):
        """
        Obtiene los campos.
        """
        if self.fields:
            return self.fields # Export fields
        else:
            if self.with_relacion: # Export fields with relation
                fields = [f.name for f in self.model._meta.get_fields() if not isinstance(f, models.UUIDField)] # Exclude UUIDField
                # remove elements before id field
                id_field = fields.index('id')
                self.fields = fields[id_field:] # Export fields with relation after id field
                return self.fields 
            else: # Export fields without relation
                self.fields = [f.name for f in self.model._meta.get_fields() if not f.is_relation and not isinstance(f, models.UUIDField)]  # Exclude UUIDField and relation
                return self.fields

    def get_custom_func_row(self, **kwargs):
        """
        Devolver una función que recibe un row y devuelve el row modificado.
        """
        def custom_func_row(row): # Custom row recibe un tupla y devuelve un tupla
            if len(kwargs) == 0:
                return row

            row_base = list(row) # Convertir la tupla a lista
            row_copy = row_base.copy() # Copiar la lista para no modificar el row original
            for key, funct in kwargs.items(): # Recorrer los kwargs
                if key in self.get_fields():
                    row_base[self.get_fields().index(key)] = funct(row_base[self.get_fields().index(key)])
                else: # Si el key no es un campo del modelo
                    # El key debe tener el formato: 'extra1_field_model' 
                    # Ejemplo: 'extra1_dia' para pasarle como parámetro a la función el valor del campo 'dia' del modelo
                    key_base = key.split('_')[1] # dia
                    value_base = row_copy[self.get_fields().index(key_base)] # Obtener el valor del campo 'dia' del modelo
                    response = funct(value_base) # Ejecutar la función con el valor del campo 'dia' del modelo
                    if isinstance(response, list):
                        row_base.extend(response)
                    else:
                        row_base.append(response)
            return tuple(row_base) # Convertir la lista a tupla
        return custom_func_row # Return una función

    def get_filename(self):
        """
        Obtiene el nombre del archivo.
        """
        if self.filename:
            return self.filename
        time_str = datetime.datetime.now().strftime('%d-%m-%Y_%H-%M')
        self.filename = self.get_sheet_name()+'{}_'.format(self.model.objects.last().id) + time_str + '.xlsx'
        return self.filename

    def get_sheet_name(self):
        """
        Obtiene el nombre de la hoja.
        """
        if self.sheet_name:
            return self.sheet_name
        try:
            self.sheet_name = self.model._meta.verbose_name_plural.title()
        except:
            try:
                self.sheet_name = self.model._meta.verbose_name.title()
            except:
                self.sheet_name = self.model.__name__
        return self.sheet_name

    def get_header(self):
        """
        Obtiene los encabezados de las columnas.
        """
        if self.header: # Si se especifica el header
            self.header = self.header
        else: # Si no se especifica el header
            self.header = [self.model._meta.get_field(field).verbose_name.title() for field in self.fields]

        if self.header_extra: # Si se especifica el header extra
            self.header.extend(self.header_extra) 
        return self.header

    def get_kwargs_custom_func_row(self):
        """
        Obtiene los kwargs para la función custom_func_row.
        """
        return {}

    def get_exception_handler(self):
        """
        Obtiene el manejador de excepciones.
        """
        if self.fields and self.header:
            if len(self.fields) != len(self.header):
                raise Exception('The number of fields and headers must be the same.')
        return super().get_exception_handler()



